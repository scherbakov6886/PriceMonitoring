#!/usr/bin/env python3

import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.stderr.reconfigure(encoding='utf-8', errors='replace')

import asyncio
import os
import re
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
import requests
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from playwright.async_api import async_playwright

# ──────────────────────────────────────────────
# НАСТРОЙКИ
# ──────────────────────────────────────────────
TELEGRAM_BOT_TOKEN = "8650471854:AAG6_uwvDZT3I5-4OuzjSzAQMCfBp7yUvcA"
TELEGRAM_CHAT_IDS  = ["435596941"]
TEMPLATE_PATH      = "Dono_мониторинг.xlsx"
OUTPUT_DIR         = "reports_dono"
HEADLESS           = True  # False — показывать браузер при отладке
# ──────────────────────────────────────────────

CITIES = ["дніпро", "павлоград"]
CITY_DISPLAY = {"дніпро": "Дніпро", "павлоград": "Павлоград"}

# Колонки шаблона (1-based)
COL_NAME   = 1
COL_DN     = 2
COL_PV     = 3
COL_WEIGHT = 4
COL_DOP_N  = 5
COL_DOP_DN = 6
COL_DOP_PV = 7
COL_SPEC_N = 8
COL_SPEC_DN= 9
COL_SPEC_PV= 10

SPEC_HEADERS = {"6. Салати", "7. Соуси", "8. Кава", "9. Напої"}

# Категорія Пиріжки виводиться в колонки A-D (як піца), а не H-J
PIROZHKY_HEADER = "5. Пиріжки"

SPECIAL_CATS_MAP = {
    "салат":   "салати",
    "соус":    "соуси",
    "кава":    "кава",
    "напо":    "напої",
    "напій":   "напої",
    "пиріж":   "пиріжки",   # → колонки A-D разом з піцою
    "пиріг":   "пиріжки",
    "пирог":   "пиріжки",
}


# ═══════════════════════════════════════════════════════════════════
# PLAYWRIGHT: вибір міста на dono.in.ua (алгоритм IQPizza)
# ═══════════════════════════════════════════════════════════════════

async def select_city(page, city: str) -> bool:
    """
    Вибір міста на dono.in.ua — той самий алгоритм що й IQPizza (EateryClub платформа):
    1. Клік «Самовивіз»
    2. Вибір міста через select або дропдаун
    3. Вибір першої адреси
    4. Підтвердження адреси
    """
    city_display = CITY_DISPLAY[city]
    print(f"\n  [1/4] Клікаю «Самовивіз»...")

    try:
        btn = page.get_by_role("button", name=re.compile(r"Самовивіз", re.IGNORECASE))
        if await btn.count() > 0:
            await btn.first.click()
            await page.wait_for_timeout(1500)
            print(f"        ✅ Клікнув «Самовивіз»")
        else:
            # Запасний варіант — шукаємо за текстом
            self_pickup = page.locator("text=Самовивіз").first
            if await self_pickup.count() > 0:
                await self_pickup.click()
                await page.wait_for_timeout(1500)
                print(f"        ✅ Клікнув «Самовивіз» (text locator)")
    except Exception as e:
        print(f"        ⚠️  Самовивіз: {e}")

    print(f"  [2/4] Вибираю місто: {city_display}...")
    selected = False

    # Спроба 1: нативний <select>
    try:
        sel = page.locator("select").first
        if await sel.count() > 0 and await sel.is_visible():
            await sel.select_option(label=city_display)
            await page.wait_for_timeout(1000)
            selected = True
            print(f"        ✅ {city_display} (через <select>)")
    except Exception:
        pass

    # Спроба 2: кастомний дропдаун (як в IQPizza)
    if not selected:
        try:
            # Відкриваємо дропдаун
            field = page.locator(
                "[class*='select__single'], [class*='select__placeholder'], "
                "[class*='select__value'], [class*='city'], "
                "[class*='Select'], [class*='dropdown']"
            ).first
            if await field.count() > 0 and await field.is_visible():
                await field.click()
            else:
                # Шукаємо за текстом "Оберіть місто" або поточне місто
                for trigger_text in ["Оберіть місто", "Дніпро", "Місто"]:
                    try:
                        trigger = page.locator(f"text={trigger_text}").first
                        if await trigger.count() > 0 and await trigger.is_visible():
                            await trigger.click()
                            break
                    except Exception:
                        pass
            await page.wait_for_timeout(800)
        except Exception as e:
            print(f"        ⚠️  Тригер дропдауну: {e}")

        # Шукаємо потрібне місто в опціях
        try:
            found = False
            for selector in [
                "[class*='select__menu'] [class*='option']",
                "[class*='select__menu-list'] > div",
                "[class*='dropdown__menu'] li",
                "[class*='menu-list'] > div",
                "[role='option']",
                "[class*='option']",
            ]:
                els = page.locator(selector)
                cnt = await els.count()
                for i in range(cnt):
                    el = els.nth(i)
                    if not await el.is_visible():
                        continue
                    if (await el.inner_text()).strip() == city_display:
                        await el.click()
                        await page.wait_for_timeout(1000)
                        found = selected = True
                        print(f"        ✅ {city_display} (через дропдаун)")
                        break
                if found:
                    break

            # XPath fallback
            if not found:
                els = page.locator(
                    f"xpath=//*[normalize-space(text())='{city_display}']"
                )
                cnt = await els.count()
                for i in range(cnt - 1, -1, -1):
                    el = els.nth(i)
                    if await el.is_visible():
                        tag = await el.evaluate("el => el.tagName.toLowerCase()")
                        if tag in ("li", "div", "span", "p", "option"):
                            await el.click()
                            await page.wait_for_timeout(1000)
                            found = selected = True
                            print(f"        ✅ {city_display} (через XPath)")
                            break
        except Exception as e:
            print(f"        ⚠️  Вибір міста: {e}")

    # Закриваємо дропдаун якщо ще відкритий
    await page.wait_for_timeout(500)
    try:
        await page.wait_for_selector(
            "[class*='select__menu'], [class*='dropdown__menu']",
            state="hidden", timeout=2000
        )
    except Exception:
        pass

    print(f"        {'✅' if selected else '⚠️ '} {city_display} "
          f"{'вибрано' if selected else 'НЕ вибрано!'}")

    print("  [3/4] Вибираю першу адресу...")
    try:
        # Чекаємо появи карток адрес
        await page.wait_for_selector(
            "[class*='pickup'], [class*='branch'], [class*='restaurant'], "
            "[class*='address-card'], [class*='shop-item'], "
            "[class*='AddressCard'], [class*='PickupPoint']",
            timeout=8000
        )
        await page.wait_for_timeout(500)

        addr_ok = False
        for sel in [
            "[class*='pickup-item']",
            "[class*='branch-item']",
            "[class*='restaurant-card']",
            "[class*='shop-card']",
            "[class*='address-item']",
            "[class*='AddressCard']",
            "[class*='PickupPoint']",
            "[class*='pickup'] li",
            "[class*='branch'] li",
        ]:
            items = page.locator(sel)
            if await items.count() > 0 and await items.first.is_visible():
                await items.first.click()
                await page.wait_for_timeout(600)
                addr_ok = True
                print(f"        ✅ Перша адреса ({sel})")
                break

        # Запасний варіант — шукаємо за назвою закладу
        if not addr_ok:
            for brand_text in ["DONO", "Dono"]:
                card = page.locator(f"text={brand_text}").first
                if await card.count() > 0 and await card.is_visible():
                    await card.click()
                    await page.wait_for_timeout(600)
                    addr_ok = True
                    print(f"        ✅ Перша адреса (за текстом '{brand_text}')")
                    break

        if not addr_ok:
            print(f"        ⚠️  Адреса не вибрана!")

    except Exception as e:
        print(f"        ⚠️  Адреса: {e}")

    print("  [4/4] Підтверджую адресу...")
    try:
        btn = page.get_by_role("button", name=re.compile(r"Підтвердити адресу", re.IGNORECASE))
        await btn.wait_for(state="visible", timeout=5000)
        await page.wait_for_timeout(400)
        await btn.click()
        await page.wait_for_timeout(3000)
        print(f"        ✅ Адресу підтверджено! URL: {page.url}")
        return True
    except Exception as e:
        print(f"        ⚠️  Підтвердження: {e}")

        # Якщо кнопка "Підтвердити" не знайдена, можливо адресу вже вибрано
        # Перевіряємо чи ми на сторінці меню
        current_url = page.url
        if any(x in current_url for x in ["menu", "catalog", "pizza", "category"]):
            print(f"        ✅ Вже на сторінці меню: {current_url}")
            return True

        return False


# ═══════════════════════════════════════════════════════════════════
# PLAYWRIGHT: модифікатори з секції «Додатки»
# ═══════════════════════════════════════════════════════════════════

def _clean_name(name: str) -> str:
    name = name.strip().rstrip('.')
    # Убираем промо-теги если они оказались в названии
    name = re.sub(r'\s*(НОВИНКА|ТОП|ЗНИЖКА|VEGETARIAN|ВЕГЕТАРІАНСЬКА|ГОСТРЕ|ХІТ)\s*$', '', name, flags=re.IGNORECASE).strip()
    return name


async def get_dodatky_from_card(page, card_locator) -> list:
    """
    Клікає по картці, знаходить секцію «Додатки» і збирає
    всі добавки з ціною.
    """
    dodatky = []
    try:
        title_el = card_locator.locator(
            '[data-pw="productCardBottomTitle"], '
            '[class*="title"], [class*="Title"], h3, h4'
        ).first
        if await title_el.count() > 0:
            await title_el.first.click()
        else:
            await card_locator.click()
        await page.wait_for_timeout(1500)

        modal_loc = page.locator(
            "[class*='modal'], [class*='Modal'], [role='dialog'], "
            "[class*='product-detail'], [class*='ProductDetail']"
        ).first
        if await modal_loc.count() == 0:
            return dodatky

        # Розкриваємо секцію «Додатки» якщо згорнута
        for section_name in ["Додатки", "Додаток", "Додатки до піци"]:
            try:
                sec = modal_loc.locator(f"text={section_name}").first
                if await sec.count() > 0:
                    await sec.click()
                    await page.wait_for_timeout(500)
                    break
            except Exception:
                pass

        modal_text = await modal_loc.inner_text()
        lines = [l.strip() for l in modal_text.split('\n') if l.strip()]

        # Известные заголовки других секций — при их появлении выходим из Додатки
        OTHER_SECTIONS = re.compile(
            r'^(бортик|модифікатор|розмір|вага|додати в кошик|замовити|склад|опис)',
            re.IGNORECASE
        )

        in_dodatky = False
        i = 0
        while i < len(lines):
            line = lines[i]
            ll   = line.lower()

            # Вход в секцию Додатки
            if re.match(r'^додатк', ll):
                in_dodatky = True
                i += 1
                continue

            if not in_dodatky:
                i += 1
                continue

            # Выход — только если встретили другой известный заголовок секции
            if OTHER_SECTIONS.match(line):
                in_dodatky = False
                i += 1
                continue

            # Цена отдельной строкой: "+ 59₴" / "+ 59" / "+59"
            pm = re.search(r'^\+\s*(\d+)', line)
            if pm:
                price = int(pm.group(1))
                name  = lines[i - 1].strip() if i > 0 else ""
                name  = re.sub(r'\s*\+\s*\d+.*$', '', name).strip()
                if name and price > 0 and not name.startswith('+') \
                        and not re.match(r'^додатк', name.lower()):
                    dodatky.append({"name": _clean_name(name), "price": price})
                i += 1
                continue

            # Цена в той же строке: "Назва + 59"
            inline = re.search(r'^(.+?)\s+\+\s*(\d+)', line)
            if inline:
                name  = inline.group(1).strip()
                price = int(inline.group(2))
                if price > 0:
                    dodatky.append({"name": _clean_name(name), "price": price})
                i += 1
                continue

            i += 1

        if not dodatky:
            i = 0
            while i < len(lines):
                line = lines[i]
                pm = re.search(r'^\+\s*(\d+)', line)
                if pm and int(pm.group(1)) > 0:
                    name = lines[i - 1].strip() if i > 0 else ""
                    name = re.sub(r'\s*\+\s*\d+.*$', '', name).strip()
                    if name and len(name) > 1 and not name.startswith('+'):
                        dodatky.append({"name": _clean_name(name), "price": int(pm.group(1))})
                else:
                    inline = re.search(r'^(.+?)\s+\+\s*(\d+)', line)
                    if inline and int(inline.group(2)) > 0:
                        dodatky.append({"name": _clean_name(inline.group(1).strip()),
                                        "price": int(inline.group(2))})
                i += 1

    except Exception as e:
        print(f"      ⚠️  Додатки: {e}")
    finally:
        try:
            close_btn = page.locator(
                "[class*='close'], [class*='Close'], "
                "[data-pw*='close'], [aria-label*='close']"
            ).first
            if await close_btn.count() > 0:
                await close_btn.click()
            else:
                await page.keyboard.press("Escape")
            await page.wait_for_timeout(500)
        except Exception:
            try:
                await page.keyboard.press("Escape")
                await page.wait_for_timeout(500)
            except Exception:
                pass

    return dodatky


# ═══════════════════════════════════════════════════════════════════
# PLAYWRIGHT: парсинг меню dono.in.ua
# ═══════════════════════════════════════════════════════════════════

async def parse_menu(page, city: str) -> dict:
    html = await page.content()
    with open(f"debug_dono_{city}.html", "w", encoding="utf-8") as f:
        f.write(html)

    result = {k: [] for k in ["піца", "пиріжки", "салати", "соуси", "кава", "напої"]}
    dodatky_collected = False

    try:
        await page.wait_for_selector(
            '[data-pw="productCard"], [class*="product-card"], '
            '[class*="ProductCard"], [class*="menu-item"]',
            timeout=15000
        )
    except Exception:
        print("  ⚠️  Картки не знайдено за 15 сек")
        return result

    cat_headers = page.locator(
        "h2.MenuCategory_title, h3.MenuCategory_title, "
        "[class*='MenuCategory'] h2, [class*='MenuCategory'] h3, "
        "[class*='category-title'], [class*='CategoryTitle']"
    )
    cat_count = await cat_headers.count()
    if cat_count == 0:
        cat_headers = page.locator("h2, h3")
        cat_count   = await cat_headers.count()

    print(f"  Знайдено категорій: {cat_count}")

    for ci in range(cat_count):
        cat_el   = cat_headers.nth(ci)
        cat_name = (await cat_el.inner_text()).strip()
        ll       = cat_name.lower()

        store_key = None
        for kw, sk in SPECIAL_CATS_MAP.items():
            if kw in ll:
                store_key = sk
                break
        if store_key is None:
            store_key = "піца"

        cat_container = cat_el.locator("xpath=ancestor::li[1]")
        if await cat_container.count() == 0:
            cat_container = cat_el.locator(
                "xpath=ancestor::*[contains(@class,'category') or "
                "contains(@class,'Category')][1]"
            )
        if await cat_container.count() == 0:
            cat_container = cat_el.locator("xpath=..")

        cards = cat_container.locator(
            '[data-pw="productCard"], [class*="product-card"], '
            '[class*="ProductCard"], [class*="menu-item"]'
        )
        card_count = await cards.count()
        print(f"  [{cat_name}] ({store_key}) — {card_count} карток")

        for idx in range(card_count):
            card = cards.nth(idx)
            try:
                name_el = card.locator(
                    '[data-pw="productCardBottomTitle"], '
                    '[class*="title"], [class*="Title"], h3, h4, h2'
                ).first
                price_el = card.locator(
                    '[data-pw="productPrice"], '
                    '[class*="price"], [class*="Price"]'
                ).first
                weight_el = card.locator(
                    '[data-pw="productWeight"], '
                    '[class*="weight"], [class*="Weight"], '
                    '[class*="gram"], [class*="volume"]'
                ).first

                if await name_el.count() == 0 or await price_el.count() == 0:
                    continue

                # Берём только первую строку — само название без описания/веса/тегов
                raw_name  = (await name_el.first.inner_text()).strip()
                name      = raw_name.split('\n')[0].strip()
                price_txt = (await price_el.first.inner_text()).strip()
                weight    = (await weight_el.first.inner_text()).strip() \
                            if await weight_el.count() > 0 else ""

                pm = re.search(r'(\d+)', price_txt.replace('\u202f', '').replace(' ', '').replace('\xa0', ''))
                if not pm:
                    continue
                price = int(pm.group(1))

                item = {
                    "name":    _clean_name(name),
                    "price":   price,
                    "weight":  weight,
                    "dodatky": [],
                }

                if store_key == "піца" and not dodatky_collected:
                    print(f"    → Збираю Додатки з: {name}")
                    dod = await get_dodatky_from_card(page, card)
                    item["dodatky"] = dod
                    dodatky_collected = True
                    if dod:
                        print(f"      ✅ {len(dod)} Додатків: {', '.join(d['name'] for d in dod[:5])}")
                    else:
                        print(f"      ⚠️  Додатки не знайдено")

                result[store_key].append(item)

            except Exception as e:
                print(f"    ⚠️  Картка {idx}: {e}")
                continue

    city_label = CITY_DISPLAY.get(city, city)
    for k, v in result.items():
        if v:
            print(f"  [{city_label}] {k}: {len(v)} позицій")

    return result


# ═══════════════════════════════════════════════════════════════════
# ШАБЛОН: парсинг категорій
# ═══════════════════════════════════════════════════════════════════

def parse_template_categories(template_path: str) -> list:
    wb  = openpyxl.load_workbook(template_path)
    ws  = wb.active
    cats = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        val_a = row[0]
        if isinstance(val_a, str) and val_a.strip():
            label = val_a.strip()
            m = re.search(r'Днепр\s+(\d+)[–\-](\d+)\s+Павлоград\s+(\d+)[–\-](\d+)', label)
            if m:
                cats.append({
                    "label":  label,
                    "dn_min": int(m.group(1)), "dn_max": int(m.group(2)),
                    "pv_min": int(m.group(3)), "pv_max": int(m.group(4)),
                })
    return cats


def get_pizza_category(price_dn: int, price_pv: int, cats: list) -> str:
    for cat in cats:
        dn_ok = cat["dn_min"] <= price_dn <= cat["dn_max"] if price_dn else False
        pv_ok = cat["pv_min"] <= price_pv <= cat["pv_max"] if price_pv else False
        if dn_ok or pv_ok:
            return cat["label"]
    return cats[-1]["label"] if cats else "Інше"


# ═══════════════════════════════════════════════════════════════════
# АРХІВ: завантаження попереднього моніторингу
# ═══════════════════════════════════════════════════════════════════

def load_archive_prices(output_dir: str, current_file: str = None) -> dict:
    if not os.path.isdir(output_dir):
        return {}

    files = sorted([
        f for f in os.listdir(output_dir)
        if f.startswith("Dono_Мониторинг_") and f.endswith(".xlsx")
        and (current_file is None or f != os.path.basename(current_file))
    ])
    if not files:
        print("  📂 Архівний файл не знайдено — порівняння недоступне")
        return {}

    latest = os.path.join(output_dir, files[-1])
    print(f"  📂 Архівний файл: {files[-1]}")

    archive = {}
    try:
        wb = openpyxl.load_workbook(latest, data_only=True)
        ws = wb.active
        sections = [
            (COL_NAME,   COL_DN,      COL_PV,     "піца"),
            (COL_NAME,   COL_DN,      COL_PV,     "пиріжки"),
            (COL_DOP_N,  COL_DOP_DN,  COL_DOP_PV, "доп"),
            (COL_SPEC_N, COL_SPEC_DN, COL_SPEC_PV,"спец"),
        ]
        for row in ws.iter_rows(min_row=2, values_only=True):
            for nc, dc, pc, key in sections:
                name = row[nc - 1]
                dn   = row[dc - 1]
                pv   = row[pc - 1]
                if not isinstance(name, str) or not name.strip():
                    continue
                k = (key, name.strip())
                existing = archive.get(k, {"dn": None, "pv": None})
                if isinstance(dn, (int, float)): existing["dn"] = dn
                if isinstance(pv, (int, float)): existing["pv"] = pv
                archive[k] = existing
        wb.close()
        print(f"  📊 Архівних позицій: {len(archive)}")
    except Exception as e:
        print(f"  ⚠️  Помилка читання архіву: {e}")
    return archive


def compare_price(new_price, arch_price):
    if not isinstance(new_price, (int, float)) or new_price == 0:
        return None
    if not isinstance(arch_price, (int, float)) or arch_price == 0:
        return None
    if new_price > arch_price: return "up"
    if new_price < arch_price: return "down"
    return None


# ═══════════════════════════════════════════════════════════════════
# EXCEL: заповнення шаблону Dono (10 колонок A-J)
# ═══════════════════════════════════════════════════════════════════

def fill_excel(all_data: dict, template_path: str) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    out_path = os.path.join(OUTPUT_DIR, f"Dono_Мониторинг_{date_str}.xlsx")
    shutil.copy(template_path, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    tmpl_cats = parse_template_categories(template_path)
    archive   = load_archive_prices(OUTPUT_DIR, current_file=out_path)

    for ri in range(ws.max_row, 1, -1):
        ws.delete_rows(ri)
    merged_ranges = [str(mc) for mc in ws.merged_cells.ranges]
    for rng in merged_ranges:
        try:
            ws.unmerge_cells(rng)
        except Exception:
            pass

    fill_cat  = PatternFill("solid", start_color="FFE5D5")
    fill_new  = PatternFill("solid", start_color="FFA500")
    fill_up   = PatternFill("solid", start_color="92D050")
    fill_down = PatternFill("solid", start_color="FF0000")
    font_cat  = Font(bold=True,  size=10)
    font_reg  = Font(bold=False, size=10)
    font_spec_hd = Font(bold=True,  size=10)
    font_spec    = Font(bold=False, size=10)
    align_c   = Alignment(horizontal="center", vertical="center", wrap_text=False)
    align_l   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    thin      = Side(style="thin",   color="CCCCCC")
    medium    = Side(style="medium", color="000000")
    brd_norm  = Border(left=thin,   right=thin,   top=thin, bottom=thin)
    brd_d     = Border(left=thin,   right=medium, top=thin, bottom=thin)
    brd_e     = Border(left=medium, right=thin,   top=thin, bottom=thin)

    col_widths = {1: 28, 2: 12, 3: 12, 4: 10, 5: 26, 6: 12, 7: 12, 8: 26, 9: 12, 10: 12}
    for col, w in col_widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    def style_row(ri: int, is_cat: bool):
        spec_val   = ws.cell(ri, COL_SPEC_N).value
        spec_is_hd = isinstance(spec_val, str) and spec_val.strip() in SPEC_HEADERS

        for ci in range(1, 11):
            cell   = ws.cell(ri, ci)
            is_num = isinstance(cell.value, (int, float))

            if ci == COL_WEIGHT:
                cell.border = brd_d
            elif ci == COL_DOP_N:
                cell.border = brd_e
            else:
                cell.border = brd_norm

            if ci <= COL_WEIGHT:
                cell.font = font_cat if is_cat else font_reg
                if is_cat:
                    cell.fill = fill_cat
                cell.alignment = align_c if is_num else align_l
            elif ci <= COL_DOP_PV:
                cell.font      = font_reg
                cell.alignment = align_c if is_num else align_l
            else:
                cell.font = font_spec_hd if spec_is_hd else font_spec
                if spec_is_hd:
                    cell.fill      = fill_cat
                    cell.alignment = align_l
                else:
                    cell.alignment = align_c if is_num else align_l

    def append_row(vals: list, is_cat: bool = False,
                   price_marks: dict = None, name_marks: set = None):
        ws.append(vals + [None] * (10 - len(vals)))
        ri = ws.max_row
        if price_marks:
            for ci, direction in price_marks.items():
                ws.cell(ri, ci).fill = fill_up if direction == "up" else fill_down
        if name_marks:
            for ci in name_marks:
                ws.cell(ri, ci).fill = fill_new
        style_row(ri, is_cat)

    def arch_marks(price_marks, name_marks, key, name, dn_val, pv_val,
                   ci_name, ci_dn, ci_pv):
        arch = archive.get((key, name))
        if arch is None:
            if ci_name:
                name_marks.add(ci_name)
        else:
            d = compare_price(dn_val or None, arch.get("dn"))
            if d: price_marks[ci_dn] = d
            d = compare_price(pv_val or None, arch.get("pv"))
            if d: price_marks[ci_pv] = d

    # База даних піц
    pizza_db = {}
    for city in CITIES:
        for item in all_data.get(city, {}).get("піца", []):
            n = item["name"]
            if n not in pizza_db:
                pizza_db[n] = {}
            pizza_db[n][city] = {
                "price":   item["price"],
                "weight":  item["weight"],
                "dodatky": item.get("dodatky", []),
            }

    def group_pizzas():
        grouped = {c["label"]: [] for c in tmpl_cats}
        for name, cities in pizza_db.items():
            dn_data  = cities.get("дніпро", {})
            pv_data  = cities.get("павлоград", {})
            dn_price = dn_data.get("price", 0) or 0
            pv_price = pv_data.get("price", 0) or 0

            lbl_dn = get_pizza_category(dn_price, 0,        tmpl_cats) if dn_price else None
            lbl_pv = get_pizza_category(0,        pv_price, tmpl_cats) if pv_price else None

            if lbl_dn == lbl_pv or (lbl_dn is None or lbl_pv is None):
                lbl = lbl_dn or lbl_pv
                grouped[lbl].append((name, cities))
            else:
                grouped[lbl_dn].append((name, {
                    "дніпро":    dn_data,
                    "павлоград": {k: v for k, v in pv_data.items() if k != "price"},
                }))
                grouped[lbl_pv].append((name, {
                    "дніпро":    {k: v for k, v in dn_data.items() if k != "price"},
                    "павлоград": pv_data,
                }))

        for lbl in grouped:
            grouped[lbl].sort(key=lambda x: (
                x[1].get("павлоград", {}).get("price", 0) or
                x[1].get("дніпро",    {}).get("price", 0) or 0
            ))
        return grouped

    grouped = group_pizzas()

    # База даних Додатки
    dop_db = {}
    for city in CITIES:
        for item in all_data.get(city, {}).get("піца", []):
            for dop in item.get("dodatky", []):
                dn = dop["name"]
                if dn not in dop_db:
                    dop_db[dn] = {}
                dop_db[dn][city] = dop["price"]

    dop_sorted = sorted(
        [(n, c) for n, c in dop_db.items() if c.get("дніпро") or c.get("павлоград")],
        key=lambda x: x[1].get("павлоград", 0) or x[1].get("дніпро", 0) or 0
    )

    # ── Пиріжки: окрема база, виводиться в A-D після піц ────────────
    pirozhky_db = {}
    for city in CITIES:
        for item in all_data.get(city, {}).get("пиріжки", []):
            n = item["name"]
            if n not in pirozhky_db:
                pirozhky_db[n] = {}
            pirozhky_db[n][city] = {
                "price":  item["price"],
                "weight": item["weight"],
            }
    pirozhky_sorted = sorted(
        pirozhky_db.items(),
        key=lambda x: (
            x[1].get("павлоград", {}).get("price", 0) or
            x[1].get("дніпро",    {}).get("price", 0) or 0
        )
    )

    # ── Спецкатегорії (H-J): Салати, Соуси, Кава, Напої ─────────────
    spec_label_map = {
        "6. Салати": "салати",
        "7. Соуси":  "соуси",
        "8. Кава":   "кава",
        "9. Напої":  "напої",
    }
    spec_stream = []
    for hdr, data_key in spec_label_map.items():
        spec_stream.append((hdr, None, None))
        spec_db_cat = {}
        for city in CITIES:
            for item in all_data.get(city, {}).get(data_key, []):
                n = item["name"]
                if n not in spec_db_cat:
                    spec_db_cat[n] = {}
                spec_db_cat[n][city] = item["price"]
        for name, cities in sorted(spec_db_cat.items(),
                                    key=lambda x: x[1].get("павлоград", 0) or x[1].get("дніпро", 0) or 0):
            dn = cities.get("дніпро",    "") or ""
            pv = cities.get("павлоград", "") or ""
            if dn != "" or pv != "":
                spec_stream.append((name, dn, pv))

    spec_idx = 0

    # ── Потік рядків: піца (A-D) → Пиріжки (A-D) ────────────────────
    # Тип рядку: "pizza_cat" | "pizza" | "pirozhky_cat" | "pirozhky"
    stream_a = []
    for cat in tmpl_cats:
        # tmpl_cats тепер містить і "5. Пиріжки" — пропускаємо його тут
        if cat["label"] == PIROZHKY_HEADER:
            continue
        lbl   = cat["label"]
        items = grouped.get(lbl, [])
        stream_a.append(("pizza_cat", lbl, None, None))
        for r in items:
            stream_a.append(("pizza", lbl, r[0], r[1]))

    # Блок Пиріжки (заголовок + позиції) в кінці A-D потоку
    stream_a.append(("pirozhky_cat", PIROZHKY_HEADER, None, None))
    for pname_p, pcities_p in pirozhky_sorted:
        stream_a.append(("pirozhky", PIROZHKY_HEADER, pname_p, pcities_p))

    dop_idx = 0

    for row_type, lbl, pname, pcities in stream_a:
        price_marks = {}
        name_marks  = set()
        is_cat = row_type in ("pizza_cat", "pirozhky_cat")

        # ── Колонки A-D ───────────────────────────────────────────────
        if row_type == "pizza" and pname and pcities:
            dn_d  = pcities.get("дніпро",   {})
            pv_d  = pcities.get("павлоград", {})
            p_dn  = dn_d.get("price", "") or ""
            p_pv  = pv_d.get("price", "") or ""
            w     = pv_d.get("weight") or dn_d.get("weight") or ""
            arch_marks(price_marks, name_marks, "піца", pname, p_dn, p_pv,
                       COL_NAME, COL_DN, COL_PV)
        elif row_type == "pirozhky" and pname and pcities:
            dn_d  = pcities.get("дніпро",   {})
            pv_d  = pcities.get("павлоград", {})
            p_dn  = dn_d.get("price", "") or ""
            p_pv  = pv_d.get("price", "") or ""
            w     = pv_d.get("weight") or dn_d.get("weight") or ""
            arch_marks(price_marks, name_marks, "пиріжки", pname, p_dn, p_pv,
                       COL_NAME, COL_DN, COL_PV)
        else:
            pname = p_dn = p_pv = w = None

        # ── Колонки E-G: Додатки ──────────────────────────────────────
        if dop_idx < len(dop_sorted):
            dop_n, dop_cities = dop_sorted[dop_idx]; dop_idx += 1
            dop_dn = dop_cities.get("дніпро",    "") or ""
            dop_pv = dop_cities.get("павлоград", "") or ""
            if not is_cat:
                arch_marks(price_marks, name_marks, "доп", dop_n, dop_dn, dop_pv,
                           COL_DOP_N, COL_DOP_DN, COL_DOP_PV)
        else:
            dop_n = dop_dn = dop_pv = None

        # ── Колонки H-J: Салати/Соуси/Кава/Напої ─────────────────────
        if spec_idx < len(spec_stream):
            s_name, s_dn, s_pv = spec_stream[spec_idx]; spec_idx += 1
            s_is_hd = s_name in SPEC_HEADERS
            if not s_is_hd and (s_dn != "" or s_pv != ""):
                arch_marks(price_marks, name_marks, "спец", s_name, s_dn, s_pv,
                           COL_SPEC_N, COL_SPEC_DN, COL_SPEC_PV)
        else:
            s_name = s_dn = s_pv = None

        if is_cat:
            append_row(
                [lbl,   None,  None,  None,
                 dop_n, dop_dn, dop_pv,
                 s_name, s_dn, s_pv],
                is_cat=True
            )
        else:
            append_row(
                [pname, p_dn, p_pv, w,
                 dop_n, dop_dn, dop_pv,
                 s_name, s_dn, s_pv],
                price_marks=price_marks, name_marks=name_marks
            )

    # ── Хвіст: залишки Додатки та Спец ───────────────────────────────
    while dop_idx < len(dop_sorted) or spec_idx < len(spec_stream):
        price_marks = {}
        name_marks  = set()

        if dop_idx < len(dop_sorted):
            dop_n, dop_cities = dop_sorted[dop_idx]; dop_idx += 1
            dop_dn = dop_cities.get("дніпро",    "") or ""
            dop_pv = dop_cities.get("павлоград", "") or ""
            arch_marks(price_marks, name_marks, "доп", dop_n, dop_dn, dop_pv,
                       COL_DOP_N, COL_DOP_DN, COL_DOP_PV)
        else:
            dop_n = dop_dn = dop_pv = None

        if spec_idx < len(spec_stream):
            s_name, s_dn, s_pv = spec_stream[spec_idx]; spec_idx += 1
            s_is_hd = s_name in SPEC_HEADERS
            if not s_is_hd and (s_dn != "" or s_pv != ""):
                arch_marks(price_marks, name_marks, "спец", s_name, s_dn, s_pv,
                           COL_SPEC_N, COL_SPEC_DN, COL_SPEC_PV)
        else:
            s_name = s_dn = s_pv = None

        if dop_n is None and s_name is None:
            break

        append_row(
            [None, None, None, None,
             dop_n, dop_dn, dop_pv,
             s_name, s_dn, s_pv],
            price_marks=price_marks, name_marks=name_marks
        )

    _add_history(wb, all_data, date_str)
    wb.save(out_path)
    print(f"  ✅ Excel збережено: {out_path}")
    return out_path


def _add_history(wb, all_data: dict, date_str: str):
    sname = "Історія"
    if sname not in wb.sheetnames:
        ws = wb.create_sheet(sname)
        ws.append(["Дата", "Місто", "Категорія", "Назва", "Ціна (грн)", "Вага"])
        for c in ws[1]:
            c.font = Font(bold=True)
    else:
        ws = wb[sname]
    city_lbl = {"дніпро": "Дніпро", "павлоград": "Павлоград"}
    cat_lbl  = {
        "піца": "Піца", "пиріжки": "Пиріжки", "салати": "Салати",
        "соуси": "Соуси", "кава": "Кава", "напої": "Напої",
    }
    for city, data in all_data.items():
        for cat_key, items in data.items():
            for item in items:
                ws.append([
                    date_str, city_lbl.get(city, city),
                    cat_lbl.get(cat_key, cat_key),
                    item.get("name", ""), item.get("price", ""), item.get("weight", ""),
                ])


# ═══════════════════════════════════════════════════════════════════
# TELEGRAM
# ═══════════════════════════════════════════════════════════════════

def send_telegram(file_path: str):
    if not TELEGRAM_BOT_TOKEN:
        print("  ⚠️  Telegram не налаштовано — пропускаю")
        return
    caption = (
        f"📊 *Моніторинг цін Dono*\n"
        f"🗓 {datetime.now().strftime('%d.%m.%Y %H:%M')}\n"
        f"🍕 Самовивіз — Дніпро + Павлоград"
    )
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"
    for chat_id in TELEGRAM_CHAT_IDS:
        try:
            with open(file_path, "rb") as f:
                r = requests.post(
                    url,
                    data={"chat_id": chat_id, "caption": caption, "parse_mode": "Markdown"},
                    files={"document": (os.path.basename(file_path), f,
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                    timeout=30,
                )
            print(f"  {'✅' if r.status_code == 200 else '❌'} Telegram → {chat_id}")
        except Exception as e:
            print(f"  ❌ {chat_id}: {e}")


# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════

async def main():
    print("=" * 52)
    print("  Dono Price Monitor v2.0 (IQPizza-style city select)")
    print(f"  {datetime.now().strftime('%d.%m.%Y  %H:%M:%S')}")
    print("=" * 52)

    if not Path(TEMPLATE_PATH).exists():
        print(f"❌ Шаблон не знайдено: {TEMPLATE_PATH}")
        return

    all_data = {}

    async with async_playwright() as p:
        for city in CITIES:
            city_label = CITY_DISPLAY.get(city, city)
            print(f"\n{'='*52}")
            print(f"  Місто: {city_label}")
            print(f"{'='*52}")

            browser = await p.chromium.launch(headless=HEADLESS)
            ctx = await browser.new_context(
                viewport={"width": 1280, "height": 900},
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
            )
            page = await ctx.new_page()

            try:
                MAX_RETRIES = 5
                RETRY_DELAY = 60
                site_ok = False

                for attempt in range(1, MAX_RETRIES + 1):
                    try:
                        # ── ГОЛОВНА ЗМІНА: завжди відкриваємо головну без slug ──
                        await page.goto(
                            "https://www.dono.in.ua/",
                            wait_until="networkidle",
                            timeout=30000,
                        )
                        await page.wait_for_timeout(2000)

                        page_text = await page.inner_text("body")

                        # Перевіряємо чи сайт доступний (не технічні роботи)
                        if any(x in page_text for x in ["In flames", "EateryClub team is working"]):
                            print(f"  ⚠️  Сайт недоступний (спроба {attempt}/{MAX_RETRIES}). "
                                  f"Чекаю {RETRY_DELAY} сек...")
                            if attempt < MAX_RETRIES:
                                await page.wait_for_timeout(RETRY_DELAY * 1000)
                                await page.reload(wait_until="networkidle", timeout=30000)
                            continue

                        if len(page_text.strip()) < 50:
                            raise Exception("Порожня сторінка")

                        site_ok = True
                        print(f"  ✅ Сайт доступний (спроба {attempt})")
                        break

                    except Exception as e:
                        print(f"  ⚠️  Помилка завантаження (спроба {attempt}/{MAX_RETRIES}): {e}")
                        if attempt < MAX_RETRIES:
                            await page.wait_for_timeout(RETRY_DELAY * 1000)

                if not site_ok:
                    print(f"  ❌ Сайт недоступний після {MAX_RETRIES} спроб. Пропускаю {city_label}.")
                    continue

                # ── ВИБІР МІСТА через алгоритм IQPizza ──────────────────────
                ok = await select_city(page, city)
                if not ok:
                    print(f"  ⚠️  Місто/адресу не підтверджено для {city_label}!")

                # Чекаємо переходу на меню
                try:
                    await page.wait_for_url(
                        re.compile(r".*(menu|catalog|pizza|category|order).*"),
                        timeout=15000
                    )
                    print(f"  ✅ URL меню: {page.url}")
                except Exception:
                    print(f"  ⚠️  URL не змінився: {page.url}")
                    # Намагаємось знайти меню через навігацію
                    for menu_text in ["Меню", "Піца", "Pizza", "Каталог"]:
                        try:
                            menu_link = page.get_by_role(
                                "link", name=re.compile(menu_text, re.IGNORECASE)
                            ).first
                            if await menu_link.count() > 0:
                                await menu_link.click()
                                await page.wait_for_timeout(2000)
                                print(f"  ✅ Перейшов через '{menu_text}': {page.url}")
                                break
                        except Exception:
                            pass

                # Чекаємо картки товарів
                try:
                    await page.wait_for_selector(
                        '[data-pw="productCard"], [class*="product-card"], '
                        '[class*="ProductCard"], [class*="menu-item"]',
                        timeout=20000
                    )
                    print("  ✅ Меню завантажено")
                except Exception:
                    print("  ⚠️  Картки не знайдено за 20 сек")

                await page.wait_for_timeout(2000)
                await page.screenshot(path=f"debug_dono_menu_{city}.png", full_page=True)

                print(f"\n  Парсю меню ({city_label})...")
                all_data[city] = await parse_menu(page, city)

            finally:
                await browser.close()
                print(f"  🔒 Браузер для {city_label} закрито")

    total = sum(len(all_data.get(c, {}).get("піца", [])) for c in CITIES)
    print(f"\n  Всього піц: {total}")
    if total == 0:
        print("  ⚠️  Дані не знайдено! Перевірте debug_dono_menu_*.png")
        return

    print("\n[3/4] Заповнюю Excel...")
    excel_path = fill_excel(all_data, TEMPLATE_PATH)

    print("\n[4/4] Відправляю в Telegram...")
    send_telegram(excel_path)

    print(f"\n✅ Готово! → {excel_path}\n")


if __name__ == "__main__":
    asyncio.run(main())
