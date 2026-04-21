#!/usr/bin/env python3

import asyncio
import os
import re
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
import requests
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from playwright.async_api import async_playwright

# ──────────────────────────────────────────────
# НАСТРОЙКИ
# ──────────────────────────────────────────────
TELEGRAM_BOT_TOKEN = ""
TELEGRAM_CHAT_IDS  = [""]
TEMPLATE_PATH      = "IQPizza_мониторинг.xlsx"
OUTPUT_DIR         = "reports"
HEADLESS           = True   # False — показывать браузер при отладке
# ──────────────────────────────────────────────

CITIES = ["дніпро", "київ"]

SITE_CATEGORY_MAP = {
    "піци":               "pizza_30",
    "піци 38 см":         "pizza_38",
    "смачно донат":       "донат_30",
    "смачно донат 38 см": "донат_38",
    "комбо":              "комбо",
    "фрітюр":             "фрітюр",
    "бокс-меню":          "бокс-меню",
}

SPECIAL_CATS = ["комбо", "фрітюр", "бокс-меню"]

# ── Фильтры для ДОП ──────────────────────────
# Суффиксы и заголовки, которые НЕ должны попадать в ДОП
_BAD_MOD_38 = re.compile(r'38\s*см', re.IGNORECASE)
_BAD_MOD_30 = re.compile(r'30\s*см', re.IGNORECASE)
# «звичайний бортик» (бесплатный) и заголовки-обёртки не включаем в ДОП
_NOT_A_MOD  = re.compile(
    r'^(без\s+цибул|звичайний\s+бортик)',
    re.IGNORECASE
)


# ═══════════════════════════════════════════════════════════════════
# PLAYWRIGHT: навигация и выбор адреса
# ═══════════════════════════════════════════════════════════════════

async def select_address(page, city: str) -> bool:
    city_display = "Київ" if city == "київ" else "Дніпро"

    print(f"  [1/4] Нажимаю «Самовывоз»...")
    try:
        btn = page.get_by_role("button", name=re.compile(r"Самовивіз", re.IGNORECASE))
        await btn.first.click()
        await page.wait_for_timeout(1500)
    except Exception as e:
        print(f"        ⚠️  {e}")

    print(f"  [2/4] Выбираю город: {city_display}...")
    selected = False

    # Вариант A: нативный <select>
    try:
        sel = page.locator("select").first
        if await sel.count() > 0:
            await sel.select_option(label=city_display)
            await page.wait_for_timeout(1000)
            selected = True
            print(f"        ✅ {city_display} (select)")
    except Exception:
        pass

    if not selected:
        # Открываем кастомный дропдаун
        try:
            field = page.locator(
                "[class*='select__single'], [class*='select__placeholder'], "
                "[class*='select__value'], [class*='city']"
            ).first
            if await field.count() > 0:
                await field.click()
            else:
                await page.locator("text=Оберіть місто").first.click()
            await page.wait_for_timeout(800)
        except Exception as e:
            print(f"        ⚠️  Триггер: {e}")

        # Ищем ТОЧНОЕ совпадение города в открытом списке
        try:
            found = False
            for selector in [
                "[class*='select__menu'] [class*='option']",
                "[class*='dropdown__menu'] li",
                "[class*='menu-list'] > div",
                "[role='option']",
            ]:
                els = page.locator(selector)
                cnt = await els.count()
                for i in range(cnt):
                    el = els.nth(i)
                    if not await el.is_visible():
                        continue
                    if (await el.inner_text()).strip() == city_display:
                        await el.click()
                        await page.wait_for_timeout(1000)
                        found = selected = True
                        break
                if found:
                    break

            if not found:
                els = page.locator(
                    f"xpath=//*[normalize-space(text())='{city_display}']"
                )
                for i in range(await els.count() - 1, -1, -1):
                    el = els.nth(i)
                    if await el.is_visible():
                        tag = await el.evaluate("el => el.tagName.toLowerCase()")
                        if tag in ("li", "div", "span", "p"):
                            await el.click()
                            await page.wait_for_timeout(1000)
                            found = selected = True
                            break
        except Exception as e:
            print(f"        ⚠️  Выбор города: {e}")

    # Проверяем что список закрылся
    await page.wait_for_timeout(500)
    try:
        await page.wait_for_selector(
            "[class*='select__menu'], [class*='dropdown__menu']",
            state="hidden", timeout=2000
        )
    except Exception:
        pass

    print(f"        {'✅' if selected else '⚠️ '} {city_display} "
          f"{'выбран' if selected else 'не выбран!'}")

    print("  [3/4] Выбираю первый адрес...")
    try:
        await page.wait_for_selector(
            "[class*='pickup'], [class*='branch'], [class*='restaurant'], "
            "[class*='address-card'], [class*='shop-item']",
            timeout=8000
        )
        await page.wait_for_timeout(500)
        addr_ok = False
        for sel in ["[class*='pickup-item']", "[class*='branch-item']",
                    "[class*='restaurant-card']", "[class*='shop-card']",
                    "[class*='address-item']"]:
            items = page.locator(sel)
            if await items.count() > 0:
                await items.first.click()
                await page.wait_for_timeout(600)
                addr_ok = True
                print(f"        ✅ Первый адрес ({sel})")
                break
        if not addr_ok:
            card = page.locator("text=IQ Pizza на").first
            if await card.count() > 0:
                await card.click()
                await page.wait_for_timeout(600)
                addr_ok = True
                print("        ✅ Первый адрес (по тексту)")
        if not addr_ok:
            print("        ⚠️  Адрес не выбран!")
    except Exception as e:
        print(f"        ⚠️  Адрес: {e}")

    print("  [4/4] Подтверждаю адрес...")
    try:
        btn = page.get_by_role("button", name=re.compile(r"Підтвердити адресу", re.IGNORECASE))
        await btn.wait_for(state="visible", timeout=5000)
        await page.wait_for_timeout(400)
        await btn.click()
        await page.wait_for_timeout(3000)
        print("        ✅ Адрес подтверждён!")
        return True
    except Exception as e:
        print(f"        ⚠️  {e}")
        return False


# ═══════════════════════════════════════════════════════════════════
# PLAYWRIGHT: получение модификаторов из карточки товара
# ═══════════════════════════════════════════════════════════════════

def _add_item(modifiers: list, section: int, name: str, price: int, size: str):
    """Добавляет модификатор или бортик в список если прошёл фильтры."""
    SECTION_BORTYK = 1
    SECTION_MOD    = 2
    if not name or price <= 0:
        return
    if section == SECTION_BORTYK:
        clean = _clean_mod_name(name)
        # Убираем повторный префикс «Бортик з» → оставляем «з сиром сулугуні»
        clean = re.sub(r'^бортик\s+', '', clean, flags=re.IGNORECASE).strip()
        # «Звичайний бортик» пропускаем (бесплатный), остальные добавляем
        if clean and not re.match(r'^звичайний', clean, re.IGNORECASE):
            modifiers.append({"name": f"Бортик {clean}", "price": price})
    elif section == SECTION_MOD:
        if _is_valid_mod(name, size):
            modifiers.append({"name": _clean_mod_name(name), "price": price})


async def get_modifiers_from_card(page, card_locator, size: str = "30") -> list:
    """
    Кликает по карточке, собирает бортики + модификаторы.
    size: "30" или "38"
    Структура карточки:
      Бортики 30 см / Бортики 38 см  ← radio-кнопки
        Звичайний бортик (бесплатный — пропускаем)
        Бортик з сиром сулугуні + 49
        Бортик з крем сиром + 59
      Подвійний сир 30 см  ˅           ← аккордеон (раскрываем)
      Модифікатори / Додатки до піці ^ ← аккордеон (раскрываем)
        Без цибулі.
        Цибуля маринована.  + 5
        ...
    Возвращает: [{"name": str, "price": int}, ...]
    """
    modifiers = []
    try:
        title_el = card_locator.locator('[data-pw="productCardBottomTitle"]')
        if await title_el.count() > 0:
            await title_el.first.click()
        else:
            await card_locator.click()
        await page.wait_for_timeout(1500)

        modal_loc = page.locator(
            "[class*='modal'], [class*='Modal'], [role='dialog']"
        ).first
        if await modal_loc.count() == 0:
            return modifiers

        # Раскрываем свёрнутые аккордеоны.
        # Для Ø30 секция называется «Модифікатори»,
        # для Ø38 — «Додатки до піці 38 см»
        accordion_texts = [
            "Модифікатори",
            f"Додатки до піці {size} см",
            f"Подвійний сир {size} см",
        ]
        for section_text in accordion_texts:
            try:
                sec = modal_loc.locator(f"text={section_text}").first
                if await sec.count() > 0:
                    await sec.click()
                    await page.wait_for_timeout(500)
            except Exception:
                pass

        modal_text = await modal_loc.inner_text()
        lines = [l.strip() for l in modal_text.split('\n') if l.strip()]

        # Секции в порядке появления в карточке:
        # NONE → BORTYK → MOD
        # Переключаемся по заголовку строки
        SECTION_NONE   = 0
        SECTION_BORTYK = 1
        SECTION_MOD    = 2
        section = SECTION_NONE

        i = 0
        while i < len(lines):
            line = lines[i]
            ll   = line.lower()

            # ── Определяем секцию ────────────────────────────────
            # Заголовок бортиков: «Бортики 30 см» или «Бортики 38 см»
            if re.match(r'^бортики\s+\d+\s*см', ll):
                section = SECTION_BORTYK
                i += 1
                continue

            # «Модифікатори» (Ø30) ИЛИ «Додатки до піці X см» (Ø38)
            if (ll.startswith('модифікатор') or
                    re.match(r'^додатки до піц', ll)):
                section = SECTION_MOD
                i += 1
                continue

            # Другие заголовки секций — сбрасываем секцию
            if (re.match(r'^подвійний\s+сир', ll) or
                re.match(r'^\d+₴', ll) or
                re.match(r'^\d+\s*г', ll)):
                section = SECTION_NONE
                i += 1
                continue

            # ── Парсим цену ──────────────────────────────────────
            # Формат 1: цена отдельной строкой «+ 49₴» или «+ 49»
            pm = re.search(r'^\+\s*(\d+)', line)
            if pm:
                price = int(pm.group(1))
                name  = lines[i - 1].strip() if i > 0 else ""
                # Убираем из имени хвост с ценой если попал
                name = re.sub(r'\s*\+\s*\d+.*$', '', name).strip()
                _add_item(modifiers, section, name, price, size)
                i += 1
                continue

            # Формат 2: «Назва + 49₴» или «Назва + 49» в одной строке
            inline = re.search(r'^(.+?)\s+\+\s*(\d+)', line)
            if inline:
                name  = inline.group(1).strip()
                price = int(inline.group(2))
                _add_item(modifiers, section, name, price, size)
                i += 1
                continue

            i += 1

    except Exception as e:
        print(f"      ⚠️  Модификаторы: {e}")
    finally:
        try:
            close_btn = page.locator(
                "[class*='close'], [class*='Close'], "
                "[data-pw*='close'], [aria-label*='close'], "
                "button[class*='modal'] svg"
            ).first
            if await close_btn.count() > 0:
                await close_btn.click()
            else:
                await page.keyboard.press("Escape")
            await page.wait_for_timeout(500)
        except Exception:
            try:
                await page.keyboard.press("Escape")
                await page.wait_for_timeout(500)
            except Exception:
                pass

    return modifiers


def _is_valid_mod(name: str, size: str) -> bool:
    """Проверяет, что модификатор подходит для данного размера и не мусор."""
    if not name or len(name) < 2:
        return False
    if _NOT_A_MOD.search(name):
        return False
    if re.match(r'^\d+$', name):
        return False
    # В ДОП Ø30 не должно быть «38 см»
    if size == "30" and _BAD_MOD_38.search(name):
        return False
    # В ДОП Ø38 не должно быть «30 см»
    if size == "38" and _BAD_MOD_30.search(name):
        return False
    return True


def _clean_mod_name(name: str) -> str:
    """Убирает суффикс размера и точку в конце."""
    name = re.sub(r'\s*\d{2}\s*см\s*$', '', name, flags=re.IGNORECASE).strip()
    return name.rstrip('.')


# ═══════════════════════════════════════════════════════════════════
# PLAYWRIGHT: парсинг меню
# ═══════════════════════════════════════════════════════════════════

async def parse_menu(page, city: str) -> dict:
    html = await page.content()
    with open(f"debug_page_{city}.html", "w", encoding="utf-8") as f:
        f.write(html)

    result = {k: [] for k in ["pizza_30", "pizza_38", "донат_30", "донат_38",
                               "комбо", "фрітюр", "бокс-меню", "соуси", "газовані напої"]}

    try:
        await page.wait_for_selector('li[data-pw="productCard"]', timeout=15000)
    except Exception:
        print("  ⚠️  Карточки не найдены за 15 сек")
        return result

    cat_headers = page.locator("h3.MenuCategory_title__PNMr7, [class*='MenuCategory'] h3")
    cat_count   = await cat_headers.count()
    if cat_count == 0:
        cat_headers = page.locator("h3")
        cat_count   = await cat_headers.count()

    print(f"  Найдено категорий: {cat_count}")

    mods_collected_30 = False
    mods_collected_38 = False

    for ci in range(cat_count):
        cat_el   = cat_headers.nth(ci)
        cat_name = (await cat_el.inner_text()).strip()
        cat_key  = cat_name.lower()

        if re.search(r'піц.*38|38.*піц', cat_key):
            store_key = "pizza_38"
        elif re.search(r'смачно.*донат.*38|донат.*38', cat_key):
            store_key = "донат_38"
        elif re.search(r'смачно.*донат|донат', cat_key) and '38' not in cat_key:
            store_key = "донат_30"
        elif re.search(r'^піц', cat_key) and '38' not in cat_key:
            store_key = "pizza_30"
        elif 'комбо' in cat_key:
            store_key = "комбо"
        elif 'фрітюр' in cat_key:
            store_key = "фрітюр"
        elif 'бокс' in cat_key:
            store_key = "бокс-меню"
        elif 'соус' in cat_key:
            store_key = "соуси"
        elif 'газован' in cat_key or 'напо' in cat_key:
            store_key = "газовані напої"
        else:
            continue

        cat_container = cat_el.locator("xpath=ancestor::li[1]")
        if await cat_container.count() == 0:
            cat_container = cat_el.locator(
                "xpath=ancestor::*[contains(@class,'category') or "
                "contains(@class,'Category')][1]"
            )
        if await cat_container.count() == 0:
            cat_container = cat_el.locator("xpath=..")

        cards      = cat_container.locator('li[data-pw="productCard"]')
        card_count = await cards.count()
        print(f"  [{cat_name}] ({store_key}) — {card_count} карточек")

        for idx in range(card_count):
            card = cards.nth(idx)
            try:
                name_el   = card.locator('[data-pw="productCardBottomTitle"]')
                price_el  = card.locator('[data-pw="productPrice"]')
                weight_el = card.locator('[data-pw="productWeight"]')

                if await name_el.count() == 0 or await price_el.count() == 0:
                    continue

                name      = (await name_el.first.inner_text()).strip()
                price_txt = (await price_el.first.inner_text()).strip()
                weight    = (await weight_el.first.inner_text()).strip() \
                            if await weight_el.count() > 0 else ""

                pm = re.search(r'(\d+)', price_txt.replace('\u202f', '').replace(' ', ''))
                if not pm:
                    continue
                price = int(pm.group(1))

                if store_key in ("pizza_30", "донат_30"):
                    clean_name = re.sub(r'\s*30\s*см', '', name).strip()
                elif store_key in ("pizza_38", "донат_38"):
                    clean_name = re.sub(r'\s*38\s*см', '', name).strip()
                else:
                    clean_name = name

                item = {
                    "name":      clean_name,
                    "orig_name": name,
                    "price":     price,
                    "weight":    weight,
                    "modifiers": [],
                }

                # Собираем модификаторы из первой карточки pizza_30
                if store_key == "pizza_30" and not mods_collected_30:
                    print(f"    → Собираю модификаторы Ø30 из: {clean_name}")
                    mods = await get_modifiers_from_card(page, card, size="30")
                    item["modifiers"] = mods
                    mods_collected_30 = True
                    if mods:
                        print(f"      ✅ {len(mods)} модификаторов Ø30: "
                              f"{', '.join(m['name'] for m in mods[:5])}")
                    else:
                        print(f"      ⚠️  Модификаторы Ø30 не найдены")

                # Собираем модификаторы из первой карточки pizza_38
                # Если первая карточка вернула пустой список — пробуем следующие (до 3х)
                elif store_key == "pizza_38" and not mods_collected_38:
                    print(f"    → Собираю модификаторы Ø38 из: {clean_name}")
                    mods = await get_modifiers_from_card(page, card, size="38")
                    if not mods and idx < card_count - 1:
                        print(f"      ⚠️  Пусто, пробую следующую карточку...")
                        # Пробуем ещё до 2 карточек
                        for retry_idx in range(idx + 1, min(idx + 3, card_count)):
                            retry_card = cards.nth(retry_idx)
                            try:
                                retry_name_el = retry_card.locator('[data-pw="productCardBottomTitle"]')
                                if await retry_name_el.count() == 0:
                                    continue
                                retry_name = (await retry_name_el.first.inner_text()).strip()
                                print(f"      → Повтор Ø38 из: {retry_name}")
                                mods = await get_modifiers_from_card(page, retry_card, size="38")
                                if mods:
                                    break
                            except Exception as e:
                                print(f"      ⚠️  Повтор {retry_idx}: {e}")
                    item["modifiers"] = mods
                    mods_collected_38 = True
                    if mods:
                        print(f"      ✅ {len(mods)} модификаторов Ø38: "
                              f"{', '.join(m['name'] for m in mods[:5])}")
                    else:
                        print(f"      ⚠️  Модификаторы Ø38 не найдены — ДОП Ø38 будет пустым!")

                result[store_key].append(item)

            except Exception as e:
                print(f"    ⚠️  Карточка {idx}: {e}")
                continue

    city_label = "Киев" if city == "київ" else "Днепр"
    for k, v in result.items():
        if v:
            print(f"  [{city_label}] {k}: {len(v)} позиций")

    return result


# ═══════════════════════════════════════════════════════════════════
# ШАБЛОН: парсинг категорий и диапазонов цен
# ═══════════════════════════════════════════════════════════════════

def parse_template_categories(template_path: str) -> dict:
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    cats_30, cats_38, special = [], [], []

    for row in ws.iter_rows(min_row=2, values_only=True):
        val_a = row[0]
        val_e = row[4]
        val_o = row[14]

        if isinstance(val_a, str) and val_a.strip():
            label = val_a.strip()
            m = re.search(r'Днепр\s+(\d+)[–\-](\d+)\s+Киев\s+(\d+)[–\-](\d+)', label)
            if m:
                cats_30.append({
                    "label":  label,
                    "dn_min": int(m.group(1)), "dn_max": int(m.group(2)),
                    "kv_min": int(m.group(3)), "kv_max": int(m.group(4)),
                })
            elif 'Донат' in label:
                cats_30.append({
                    "label": label, "is_donat": True,
                    "dn_min": 9000, "dn_max": 9999,
                    "kv_min": 9000, "kv_max": 9999,
                })

        if isinstance(val_e, str) and val_e.strip():
            label = val_e.strip()
            m = re.search(r'Днепр\s+(\d+)[–\-](\d+)\s+Киев\s+(\d+)[–\-](\d+)', label)
            if m:
                cats_38.append({
                    "label":  label,
                    "dn_min": int(m.group(1)), "dn_max": int(m.group(2)),
                    "kv_min": int(m.group(3)), "kv_max": int(m.group(4)),
                })
            elif 'Донат' in label:
                cats_38.append({
                    "label": label, "is_donat": True,
                    "dn_min": 9000, "dn_max": 9999,
                    "kv_min": 9000, "kv_max": 9999,
                })

        if isinstance(val_o, str) and val_o.strip():
            special.append(val_o.strip())

    return {"pizza_30": cats_30, "pizza_38": cats_38, "special": special}


def get_pizza_category(price_dn: int, price_kv: int, cats: list) -> str:
    """
    Определяет категорию по цене.
    Вызывается либо с одной ценой (вторая = 0), либо с обеими.
    """
    for cat in cats:
        if cat.get("is_donat"):
            continue
        dn_ok = cat["dn_min"] <= price_dn <= cat["dn_max"] if price_dn else False
        kv_ok = cat["kv_min"] <= price_kv <= cat["kv_max"] if price_kv else False
        if dn_ok or kv_ok:
            return cat["label"]
    for cat in cats:
        if cat.get("is_donat"):
            return cat["label"]
    return cats[-1]["label"] if cats else "Інше"


# ═══════════════════════════════════════════════════════════════════
# АРХИВ: загрузка предыдущего мониторинга для сравнения цен
# ═══════════════════════════════════════════════════════════════════

def load_archive_prices(output_dir: str, current_file: str = None) -> dict:
    """
    Ищет самый свежий файл IQPizza_Мониторинг_*.xlsx в output_dir,
    исключая current_file (текущий сохраняемый файл).

    Одна позиция может встречаться в двух строках (dn в одной, kv в другой)
    из-за split-логики категорий — мержим по имени, не перезаписываем.

    Возвращает:
      {(key, name): {"dn": price_or_None, "kv": price_or_None}}
    key: "30" | "38" | "доп30" | "доп38" | "спец"
    """
    if not os.path.isdir(output_dir):
        return {}

    files = sorted([
        f for f in os.listdir(output_dir)
        if f.startswith("IQPizza_Мониторинг_") and f.endswith(".xlsx")
        and (current_file is None or f != os.path.basename(current_file))
    ])
    if not files:
        print("  📂 Архивный файл не найден — сравнение недоступно")
        return {}

    # Берём самый свежий (последний в отсортированном списке)
    latest = os.path.join(output_dir, files[-1])
    print(f"  📂 Архивный файл: {files[-1]}")

    archive = {}
    try:
        wb = openpyxl.load_workbook(latest, data_only=True)
        ws = wb.active
        # (name_col, dn_col, kv_col, key)
        sections = [
            (1,  2,  3,  "30"),
            (5,  6,  7,  "38"),
            (9,  10, 11, "доп30"),
            (12, 13, 14, "доп38"),
            (15, 16, 17, "спец"),
        ]
        for row in ws.iter_rows(min_row=2, values_only=True):
            for nc, dc, kc, key in sections:
                name = row[nc - 1]
                dn   = row[dc - 1]
                kv   = row[kc - 1]
                if not isinstance(name, str) or not name.strip():
                    continue
                k = (key, name.strip())
                existing = archive.get(k, {"dn": None, "kv": None})
                # Мержим: не перезаписываем уже найденную цену нулём/None
                if isinstance(dn, (int, float)):
                    existing["dn"] = dn
                if isinstance(kv, (int, float)):
                    existing["kv"] = kv
                archive[k] = existing
        wb.close()
        print(f"  📊 Архивных позиций: {len(archive)}")
    except Exception as e:
        print(f"  ⚠️  Ошибка чтения архива: {e}")
    return archive


def compare_price(new_price, arch_price):
    """Возвращает 'up'|'down'|None."""
    if not isinstance(new_price, (int, float)) or new_price == 0:
        return None
    if not isinstance(arch_price, (int, float)) or arch_price == 0:
        return None
    if new_price > arch_price:
        return "up"
    if new_price < arch_price:
        return "down"
    return None


# ═══════════════════════════════════════════════════════════════════
# EXCEL: заполнение шаблона
# ═══════════════════════════════════════════════════════════════════

def fill_excel(all_data: dict, template_path: str) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    out_path = os.path.join(OUTPUT_DIR, f"IQPizza_Мониторинг_{date_str}.xlsx")
    shutil.copy(template_path, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    tmpl_cats = parse_template_categories(template_path)

    # Загружаем архивные цены для сравнения (исключаем текущий файл)
    archive = load_archive_prices(OUTPUT_DIR, current_file=out_path)

    for ri in range(ws.max_row, 1, -1):
        ws.delete_rows(ri)

    # Снимаем все объединённые ячейки из шаблона — они мешают записи данных
    merged_ranges = [str(mc) for mc in ws.merged_cells.ranges]
    for rng in merged_ranges:
        try:
            ws.unmerge_cells(rng)
        except Exception:
            pass

    # ── Стили ─────────────────────────────────────────────────────
    fill_cat     = PatternFill("solid", start_color="FFE5D5")  # заголовок категории
    fill_new     = PatternFill("solid", start_color="FFA500")  # новая номенклатура
    fill_up      = PatternFill("solid", start_color="92D050")  # цена выросла  → зелёный
    fill_down    = PatternFill("solid", start_color="FF0000")  # цена упала    → красный
    font_cat     = Font(bold=True,  size=10)
    font_spec_hd = Font(bold=True,  size=10)
    font_spec    = Font(bold=False, size=10)
    font_reg     = Font(size=10)
    align_c      = Alignment(horizontal="center", vertical="center", wrap_text=False)
    align_l      = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    thin         = Side(style="thin",   color="CCCCCC")
    medium       = Side(style="medium", color="000000")
    # Вертикальная разделительная линия между col8 (H, Вес38) и col9 (I, ДОП Ø30)
    brd_normal   = Border(left=thin,   right=thin,   top=thin, bottom=thin)
    brd_h        = Border(left=thin,   right=medium, top=thin, bottom=thin)  # col 8  — жирная правая
    brd_i        = Border(left=medium, right=thin,   top=thin, bottom=thin)  # col 9  — жирная левая
    brd_n        = brd_normal
    brd_o        = brd_normal

    SPEC_HEADERS = {"6. КОМБО", "7. Фрітюр", "8. Бокс-меню", "9. Соуси", "10. Газовані напої"}

    col_widths = {
        1:32, 2:10, 3:10, 4:10,
        5:32, 6:10, 7:10, 8:10,
        9:28, 10:10, 11:10,
        12:28, 13:10, 14:10,
        15:28, 16:10, 17:10,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    def style_row(ri: int, is_cat: bool):
        """
        Заливка фоном — только колонки 1-14 (до вертикальной линии).
        Колонки 15-17 — без заливки (кроме заголовков спецкатегорий).
        Вертикальная разделительная линия: правая граница col14 + левая граница col15 — medium.
        """
        spec_val      = ws.cell(row=ri, column=15).value
        spec_is_hd    = isinstance(spec_val, str) and spec_val.strip() in SPEC_HEADERS

        for ci in range(1, 18):
            cell = ws.cell(row=ri, column=ci)
            # Граница — вертикальная разделительная линия между col8 и col9
            if ci == 8:
                cell.border = brd_h
            elif ci == 9:
                cell.border = brd_i
            else:
                cell.border = brd_normal

            # Шрифт и выравнивание
            is_num = isinstance(cell.value, (int, float))
            if ci <= 8:
                # Пиццы Ø30 и Ø38 (A-H): заливка fill_cat для заголовков
                cell.font = font_cat if is_cat else font_reg
                if is_cat:
                    cell.fill = fill_cat
                cell.alignment = align_c if is_num else align_l
            elif ci <= 14:
                # ДОП Ø30/Ø38 (I-N): никогда жирный, никогда заливка фоном
                cell.font = font_reg
                # fill НЕ трогаем — mark_cell мог уже установить цвет
                cell.alignment = align_c if is_num else align_l
            else:
                # Спецкатегории O-Q: жирный только для заголовков (6-10)
                cell.font = font_spec_hd if spec_is_hd else font_spec
                if spec_is_hd:
                    cell.fill      = fill_cat
                    cell.alignment = align_l
                else:
                    # fill НЕ трогаем — mark_cell мог уже установить цвет
                    cell.alignment = align_c if is_num else align_l

    def mark_cell(ri: int, ci: int, fill: PatternFill):
        """Устанавливает заливку конкретной ячейке (для сравнения цен)."""
        ws.cell(row=ri, column=ci).fill = fill

    def append_row(vals: list, is_cat: bool = False,
                   price_marks: dict = None, name_marks: set = None):
        """
        vals        — 17 значений
        is_cat      — строка-заголовок категории
        price_marks — {col_index: "up"|"down"} для раскраски цен
        name_marks  — {col_index} колонки с новой номенклатурой (оранжевый)
        """
        ws.append(vals + [None] * (17 - len(vals)))
        ri = ws.max_row
        # Сначала применяем цветовые метки (до style_row, чтобы style_row не перетёр)
        if price_marks:
            for ci, direction in price_marks.items():
                ws.cell(row=ri, column=ci).fill = fill_up if direction == "up" else fill_down
        if name_marks:
            for ci in name_marks:
                ws.cell(row=ri, column=ci).fill = fill_new
        style_row(ri, is_cat)

    # ── Базы данных пицц ──────────────────────────────────────────
    def build_pizza_db(size_key: str) -> dict:
        db = {}
        for city in CITIES:
            for item in all_data.get(city, {}).get(size_key, []):
                n = item["name"]
                if n not in db:
                    db[n] = {}
                db[n][city] = {
                    "price":     item["price"],
                    "weight":    item["weight"],
                    "modifiers": item.get("modifiers", []),
                }
        return db

    pizza30_db = build_pizza_db("pizza_30")
    pizza38_db = build_pizza_db("pizza_38")

    for city in CITIES:
        for item in all_data.get(city, {}).get("донат_30", []):
            n = item["name"]
            if n not in pizza30_db:
                pizza30_db[n] = {}
            pizza30_db[n][city] = {
                "price": item["price"], "weight": item["weight"],
                "modifiers": item.get("modifiers", []), "is_donat": True,
            }
        for item in all_data.get(city, {}).get("донат_38", []):
            n = item["name"]
            if n not in pizza38_db:
                pizza38_db[n] = {}
            pizza38_db[n][city] = {
                "price": item["price"], "weight": item["weight"],
                "modifiers": item.get("modifiers", []), "is_donat": True,
            }

    def group_pizzas(db: dict, cats: list) -> dict:
        """
        Группирует пиццы по ценовым категориям.

        Если цена Днепра и цена Киева попадают в РАЗНЫЕ категории —
        позиция вставляется ДВАЖДЫ:
          • в категорию Днепра — с ценой Днепра и пустой ценой Киева
          • в категорию Киева  — с пустой ценой Днепра и ценой Киева
        """
        grouped = {c["label"]: [] for c in cats}

        for name, cities in db.items():
            dn_data  = cities.get("дніпро", {})
            kv_data  = cities.get("київ",   {})
            dn_price = dn_data.get("price", 0) or 0
            kv_price = kv_data.get("price", 0) or 0
            is_donat = any(v.get("is_donat") for v in cities.values())

            if is_donat:
                lbl = next((c["label"] for c in cats if c.get("is_donat")), cats[-1]["label"])
                grouped[lbl].append((name, cities))
                continue

            lbl_dn = get_pizza_category(dn_price, 0,        cats) if dn_price else None
            lbl_kv = get_pizza_category(0,        kv_price, cats) if kv_price else None

            if lbl_dn == lbl_kv or (lbl_dn is None or lbl_kv is None):
                # Обе цены в одной категории — одна строка, обе цены
                lbl = lbl_dn or lbl_kv
                grouped[lbl].append((name, cities))
            else:
                # Цены в РАЗНЫХ категориях — две строки с пустой ценой другого города
                cities_dn_only = {
                    "дніпро": dn_data,
                    # Киев — только вес, цена пустая
                    "київ": {k: v for k, v in kv_data.items() if k != "price"},
                }
                cities_kv_only = {
                    # Днепр — только вес, цена пустая
                    "дніпро": {k: v for k, v in dn_data.items() if k != "price"},
                    "київ":   kv_data,
                }
                grouped[lbl_dn].append((name, cities_dn_only))
                grouped[lbl_kv].append((name, cities_kv_only))

        for lbl in grouped:
            grouped[lbl].sort(key=lambda x: (
                x[1].get("київ",   {}).get("price", 0) or
                x[1].get("дніпро", {}).get("price", 0) or 0
            ))
        return grouped

    grouped30 = group_pizzas(pizza30_db, tmpl_cats["pizza_30"])
    grouped38 = group_pizzas(pizza38_db, tmpl_cats["pizza_38"])

    # ── Базы данных модификаторов ─────────────────────────────────
    # build_mod_db собирает ДОП для обоих размеров,
    # включая бортики (они хранятся в modifiers первой карточки соответствующей категории)
    def build_mod_db(size_key: str, size: str) -> dict:
        db = {}
        for city in CITIES:
            for item in all_data.get(city, {}).get(size_key, []):
                for mod in item.get("modifiers", []):
                    mn = mod["name"]
                    # Бортики (начинаются с "Бортик ") проходят без _is_valid_mod,
                    # обычные модификаторы — через фильтр размера
                    is_bortyk = mn.lower().startswith("бортик ")
                    if not is_bortyk and not _is_valid_mod(mn, size):
                        continue
                    if mn not in db:
                        db[mn] = {}
                    db[mn][city] = mod["price"]
        return db

    mod_db_30 = build_mod_db("pizza_30", "30")
    mod_db_38 = build_mod_db("pizza_38", "38")

    mods30_sorted = sorted(mod_db_30.items(),
                           key=lambda x: x[1].get("київ", 0) or x[1].get("дніпро", 0) or 0)
    mods38_sorted = sorted(mod_db_38.items(),
                           key=lambda x: x[1].get("київ", 0) or x[1].get("дніпро", 0) or 0)

    # ── Спецкатегории (КОМБО / Фрітюр / Бокс-меню / Соуси / Напої) ─
    spec_map = {
        "6. КОМБО":           "комбо",
        "7. Фрітюр":          "фрітюр",
        "8. Бокс-меню":       "бокс-меню",
        "9. Соуси":           "соуси",
        "10. Газовані напої": "газовані напої",
    }
    spec_db = {}
    for tmpl_label, data_key in spec_map.items():
        db = {}
        for city in CITIES:
            for item in all_data.get(city, {}).get(data_key, []):
                n = item["name"]
                if n not in db:
                    db[n] = {}
                db[n][city] = {"price": item["price"], "weight": item["weight"]}
        spec_db[tmpl_label] = db

    # Два отдельных потока для спецкатегорий:
    # spec_headers_q — очередь заголовков (6.КОМБО, 7.Фрітюр...) для строк-заголовков пицц
    # spec_data_q   — очередь позиций данных для строк-данных пицц и хвоста
    spec_headers_q = []
    spec_data_q    = []
    for lbl in ["6. КОМБО", "7. Фрітюр", "8. Бокс-меню", "9. Соуси", "10. Газовані напої"]:
        spec_headers_q.append(lbl)
        for name, cities in sorted(
            spec_db.get(lbl, {}).items(),
            key=lambda x: x[1].get("київ", {}).get("price", 0) or
                          x[1].get("дніпро", {}).get("price", 0) or 0
        ):
            dn = cities.get("дніпро", {}).get("price", "") or ""
            kv = cities.get("київ",   {}).get("price", "") or ""
            if dn != "" or kv != "":   # пропускаем позиции без цен
                spec_data_q.append((name, dn, kv))

    # ── Записываем строки ─────────────────────────────────────────
    # Поток A (A-H): пиццы — заголовки категорий + строки данных.
    # Поток B (I-N): ДОП — непрерывно, на КАЖДОЙ строке (и заголовках, и данных).
    # Поток C (O-Q): спец — заголовки только на строках-заголовках пицц,
    #                        данные только на строках-данных пицц и в хвосте.

    cats30_labels = [c["label"] for c in tmpl_cats["pizza_30"]]
    cats38_labels = [c["label"] for c in tmpl_cats["pizza_38"]]

    # Строим поток A
    stream_a = []
    for cat30_lbl, cat38_lbl in zip(cats30_labels, cats38_labels):
        items30 = grouped30.get(cat30_lbl, [])
        items38 = grouped38.get(cat38_lbl, [])
        stream_a.append((True, cat30_lbl, cat38_lbl, None, None))
        max_len = max(len(items30), len(items38), 0)
        for i in range(max_len):
            r30 = items30[i] if i < len(items30) else None
            r38 = items38[i] if i < len(items38) else None
            if r30 is None and r38 is None:
                continue
            stream_a.append((False, cat30_lbl, cat38_lbl, r30, r38))

    # Поток B: ДОП (только непустые)
    mods30_list_flat = [(n, c) for n, c in mods30_sorted
                        if c.get("дніпро") or c.get("київ")]
    mods38_list_flat = [(n, c) for n, c in mods38_sorted
                        if c.get("дніпро") or c.get("київ")]
    mod30_idx = 0
    mod38_idx = 0

    # Поток C: spec_headers_q и spec_data_q уже построены выше
    spec_hdr_idx  = 0
    spec_data_idx = 0

    def _arch_marks(price_marks, name_marks, arch_key, name, dn_val, kv_val,
                    ci_name, ci_dn, ci_kv):
        arch = archive.get((arch_key, name))
        if arch is None:
            if ci_name:
                name_marks.add(ci_name)
        else:
            d = compare_price(dn_val or None, arch["dn"])
            if d: price_marks[ci_dn] = d
            d = compare_price(kv_val or None, arch["kv"])
            if d: price_marks[ci_kv] = d

    for is_header, cat30_lbl, cat38_lbl, r30, r38 in stream_a:
        price_marks = {}
        name_marks  = set()

        # Пицца Ø30
        if r30:
            n30, c30 = r30
            dn30_d = c30.get("дніпро", {})
            kv30_d = c30.get("київ",   {})
            p_dn30 = dn30_d.get("price", "") or ""
            p_kv30 = kv30_d.get("price", "") or ""
            w30    = kv30_d.get("weight") or dn30_d.get("weight") or ""
            _arch_marks(price_marks, name_marks, "30", n30, p_dn30, p_kv30, 1, 2, 3)
        else:
            n30 = p_dn30 = p_kv30 = w30 = None

        # Пицца Ø38
        if r38:
            n38, c38 = r38
            dn38_d = c38.get("дніпро", {})
            kv38_d = c38.get("київ",   {})
            p_dn38 = dn38_d.get("price", "") or ""
            p_kv38 = kv38_d.get("price", "") or ""
            w38    = kv38_d.get("weight") or dn38_d.get("weight") or ""
            _arch_marks(price_marks, name_marks, "38", n38, p_dn38, p_kv38, 5, 6, 7)
        else:
            n38 = p_dn38 = p_kv38 = w38 = None

        # ДОП Ø30 — на КАЖДОЙ строке
        if mod30_idx < len(mods30_list_flat):
            d30_n, d30_cities = mods30_list_flat[mod30_idx]; mod30_idx += 1
            d30_dn = d30_cities.get("дніпро", "") or ""
            d30_kv = d30_cities.get("київ",   "") or ""
            if not is_header:
                _arch_marks(price_marks, name_marks, "доп30", d30_n, d30_dn, d30_kv, 9, 10, 11)
        else:
            d30_n = d30_dn = d30_kv = None

        # ДОП Ø38 — на КАЖДОЙ строке
        if mod38_idx < len(mods38_list_flat):
            d38_n, d38_cities = mods38_list_flat[mod38_idx]; mod38_idx += 1
            d38_dn = d38_cities.get("дніпро", "") or ""
            d38_kv = d38_cities.get("київ",   "") or ""
            if not is_header:
                _arch_marks(price_marks, name_marks, "доп38", d38_n, d38_dn, d38_kv, 12, 13, 14)
        else:
            d38_n = d38_dn = d38_kv = None

        # Спец O-Q
        if is_header:
            # Берём следующий заголовок спецкатегории
            if spec_hdr_idx < len(spec_headers_q):
                s_name = spec_headers_q[spec_hdr_idx]; spec_hdr_idx += 1
                s_dn = s_kv = None
            else:
                s_name = s_dn = s_kv = None
        else:
            # Берём следующую позицию данных
            if spec_data_idx < len(spec_data_q):
                s_name, s_dn, s_kv = spec_data_q[spec_data_idx]; spec_data_idx += 1
                _arch_marks(price_marks, name_marks, "спец", s_name, s_dn, s_kv, 15, 16, 17)
            else:
                s_name = s_dn = s_kv = None

        if is_header:
            append_row(
                [cat30_lbl, None, None, None,
                 cat38_lbl, None, None, None,
                 d30_n, d30_dn, d30_kv,
                 d38_n, d38_dn, d38_kv,
                 s_name, s_dn, s_kv],
                is_cat=True
            )
        else:
            append_row([
                n30, p_dn30, p_kv30, w30,
                n38, p_dn38, p_kv38, w38,
                d30_n, d30_dn, d30_kv,
                d38_n, d38_dn, d38_kv,
                s_name, s_dn, s_kv,
            ], price_marks=price_marks, name_marks=name_marks)

    # ── Хвост: оставшиеся ДОП и спец ─────────────────────────────
    while (mod30_idx < len(mods30_list_flat) or
           mod38_idx < len(mods38_list_flat) or
           spec_data_idx < len(spec_data_q)):
        price_marks = {}
        name_marks  = set()

        if mod30_idx < len(mods30_list_flat):
            d30_n, d30_cities = mods30_list_flat[mod30_idx]; mod30_idx += 1
            d30_dn = d30_cities.get("дніпро", "") or ""
            d30_kv = d30_cities.get("київ",   "") or ""
            _arch_marks(price_marks, name_marks, "доп30", d30_n, d30_dn, d30_kv, 9, 10, 11)
        else:
            d30_n = d30_dn = d30_kv = None

        if mod38_idx < len(mods38_list_flat):
            d38_n, d38_cities = mods38_list_flat[mod38_idx]; mod38_idx += 1
            d38_dn = d38_cities.get("дніпро", "") or ""
            d38_kv = d38_cities.get("київ",   "") or ""
            _arch_marks(price_marks, name_marks, "доп38", d38_n, d38_dn, d38_kv, 12, 13, 14)
        else:
            d38_n = d38_dn = d38_kv = None

        if spec_data_idx < len(spec_data_q):
            s_name, s_dn, s_kv = spec_data_q[spec_data_idx]; spec_data_idx += 1
            _arch_marks(price_marks, name_marks, "спец", s_name, s_dn, s_kv, 15, 16, 17)
        else:
            s_name = s_dn = s_kv = None

        # Пропускаем строку если всё пустое
        if d30_n is None and d38_n is None and s_name is None:
            break

        append_row([
            None, None, None, None,
            None, None, None, None,
            d30_n, d30_dn, d30_kv,
            d38_n, d38_dn, d38_kv,
            s_name, s_dn, s_kv,
        ], price_marks=price_marks, name_marks=name_marks)

    _add_history(wb, all_data, date_str)
    wb.save(out_path)
    print(f"  ✅ Excel сохранён: {out_path}")
    return out_path


def _add_history(wb, all_data: dict, date_str: str):
    sname = "Історія"
    if sname not in wb.sheetnames:
        ws = wb.create_sheet(sname)
        ws.append(["Дата", "Город", "Категория", "Название", "Цена (грн)", "Вес"])
        for c in ws[1]:
            c.font = Font(bold=True)
    else:
        ws = wb[sname]
    city_lbl = {"київ": "Киев", "дніпро": "Днепр"}
    cat_lbl  = {
        "pizza_30": "Ø30 см", "pizza_38": "Ø38 см",
        "донат_30": "Донат 30", "донат_38": "Донат 38",
        "комбо": "КОМБО", "фрітюр": "Фрітюр", "бокс-меню": "Бокс-меню",
        "соуси": "Соусы", "газовані напої": "Газированные напитки",
    }
    for city, data in all_data.items():
        for cat_key, items in data.items():
            for item in items:
                ws.append([
                    date_str, city_lbl.get(city, city),
                    cat_lbl.get(cat_key, cat_key),
                    item.get("name", ""), item.get("price", ""), item.get("weight", ""),
                ])


# ═══════════════════════════════════════════════════════════════════
# TELEGRAM
# ═══════════════════════════════════════════════════════════════════

def send_telegram(file_path: str):
    if not TELEGRAM_BOT_TOKEN:
        print("  ⚠️  Telegram не настроен — пропускаю")
        return
    caption = (
        f"📊 *Мониторинг цен IQ Pizza*\n"
        f"🗓 {datetime.now().strftime('%d.%m.%Y %H:%M')}\n"
        f"🍕 Самовывоз — Киев + Днепр"
    )
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"
    for chat_id in TELEGRAM_CHAT_IDS:
        try:
            with open(file_path, "rb") as f:
                r = requests.post(
                    url,
                    data={"chat_id": chat_id, "caption": caption, "parse_mode": "Markdown"},
                    files={"document": (os.path.basename(file_path), f,
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                    timeout=30,
                )
            print(f"  {'✅' if r.status_code == 200 else '❌'} Telegram → {chat_id}")
        except Exception as e:
            print(f"  ❌ {chat_id}: {e}")


# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════

async def main():
    print("=" * 52)
    print("  IQ Pizza Price Monitor v3.1")
    print(f"  {datetime.now().strftime('%d.%m.%Y  %H:%M:%S')}")
    print("=" * 52)

    if not Path(TEMPLATE_PATH).exists():
        print(f"❌ Шаблон не найден: {TEMPLATE_PATH}")
        return

    all_data = {}

    async with async_playwright() as p:
        for city in CITIES:
            city_label = "Киев" if city == "київ" else "Днепр"
            print(f"\n{'='*52}")
            print(f"  Город: {city_label}")
            print(f"{'='*52}")

            browser = await p.chromium.launch(headless=HEADLESS)
            ctx = await browser.new_context(
                viewport={"width": 1280, "height": 900},
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
            )
            page = await ctx.new_page()

            try:
                MAX_RETRIES = 5
                RETRY_DELAY = 60
                site_ok = False

                for attempt in range(1, MAX_RETRIES + 1):
                    try:
                        await page.goto(
                            "https://iqpizza.com.ua/address",
                            wait_until="networkidle",
                            timeout=30000,
                        )
                        await page.wait_for_timeout(2000)

                        page_text = await page.inner_text("body")
                        if any(x in page_text for x in ["In flames", "EateryClub team is working"]):
                            print(f"  ⚠️  Сайт недоступен (попытка {attempt}/{MAX_RETRIES}). "
                                  f"Жду {RETRY_DELAY} сек...")
                            if attempt < MAX_RETRIES:
                                await page.wait_for_timeout(RETRY_DELAY * 1000)
                                await page.reload(wait_until="networkidle", timeout=30000)
                            continue

                        site_ok = True
                        print(f"  ✅ Сайт доступен (попытка {attempt})")
                        break

                    except Exception as e:
                        print(f"  ⚠️  Ошибка загрузки (попытка {attempt}/{MAX_RETRIES}): {e}")
                        if attempt < MAX_RETRIES:
                            await page.wait_for_timeout(RETRY_DELAY * 1000)

                if not site_ok:
                    print(f"  ❌ Сайт недоступен после {MAX_RETRIES} попыток. Пропускаю {city_label}.")
                    continue

                ok = await select_address(page, city=city)
                if not ok:
                    print(f"  ⚠️  Адрес не подтверждён для {city_label}!")

                try:
                    await page.wait_for_url(
                        re.compile(r".*(menu|category|pizza|catalog|restaurants).*"),
                        timeout=15000
                    )
                    print(f"  ✅ URL: {page.url}")
                except Exception:
                    print(f"  ⚠️  URL не изменился: {page.url}")
                    print("  → Перехожу в меню напрямую...")
                    await page.goto("https://iqpizza.com.ua/menu", wait_until="networkidle")
                    await page.wait_for_timeout(2000)
                    print(f"  ✅ URL: {page.url}")

                try:
                    await page.wait_for_selector('li[data-pw="productCard"]', timeout=20000)
                    print("  ✅ Меню загружено")
                except Exception:
                    print("  ⚠️  Карточки не найдены за 20 сек")

                await page.wait_for_timeout(2000)
                await page.screenshot(path=f"debug_menu_{city}.png", full_page=True)

                print(f"\n  Парсю меню ({city_label})...")
                all_data[city] = await parse_menu(page, city)

            finally:
                await browser.close()
                print(f"  🔒 Браузер для {city_label} закрыт")

    total_pizzas = sum(
        len(all_data.get(c, {}).get("pizza_30", [])) +
        len(all_data.get(c, {}).get("pizza_38", []))
        for c in CITIES
    )
    print(f"\n  Всего пицц: {total_pizzas}")
    if total_pizzas == 0:
        print("  ⚠️  Данные не найдены! Проверьте debug_menu_*.png")
        return

    print("\n[3/4] Заполняю Excel...")
    excel_path = fill_excel(all_data, TEMPLATE_PATH)

    print("\n[4/4] Отправляю в Telegram...")
    send_telegram(excel_path)

    print(f"\n✅ Готово! → {excel_path}\n")


if __name__ == "__main__":
    asyncio.run(main())
